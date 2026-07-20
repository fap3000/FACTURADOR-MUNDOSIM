"""
Mundosim Facturador - App Web
Backend Flask con SQLite para gestion de comprobantes ARCA
Soporta Factura A (tipo 1) y Factura B (tipo 6)
"""

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash
from functools import wraps
import sqlite3
import json
import os
import csv
import io
from datetime import datetime

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "mundosim-secret-2024")

DB_PATH = os.environ.get("DB_PATH", "facturador.db")

USUARIOS = {
    "mendoza": {
        "password": os.environ.get("PASS_MENDOZA", "mendoza123"),
        "sucursal": "Mendoza",
        "punto_venta": int(os.environ.get("PV_MENDOZA", "1")),
    },
    "sanjuan": {
        "password": os.environ.get("PASS_SANJUAN", "sanjuan123"),
        "sucursal": "San Juan",
        "punto_venta": int(os.environ.get("PV_SANJUAN", "2")),
    },
    "admin": {
        "password": os.environ.get("PASS_ADMIN", "admin2024"),
        "sucursal": "Administracion",
        "punto_venta": None,
    },
}

TIPOS_COMPROBANTE = {
    "A": {"codigo": 1,  "nombre": "Factura A", "iva": True},
    "B": {"codigo": 6,  "nombre": "Factura B", "iva": False},
}

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS comprobantes (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                numero       INTEGER,
                tipo         TEXT,
                punto_venta  INTEGER,
                sucursal     TEXT,
                fecha        TEXT,
                destinatario TEXT,
                doc_tipo     INTEGER,
                doc_nro      TEXT,
                neto         REAL,
                iva          REAL,
                monto        REAL,
                cae          TEXT,
                cae_vto      TEXT,
                estado       TEXT,
                error        TEXT,
                usuario      TEXT,
                created_at   TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()

init_db()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        usuario = request.form.get("usuario", "").lower()
        password = request.form.get("password", "")
        if usuario in USUARIOS and USUARIOS[usuario]["password"] == password:
            session["usuario"] = usuario
            session["sucursal"] = USUARIOS[usuario]["sucursal"]
            session["punto_venta"] = USUARIOS[usuario]["punto_venta"]
            return redirect(url_for("dashboard"))
        error = "Usuario o contrasena incorrectos"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def dashboard():
    with get_db() as conn:
        if session["usuario"] == "admin":
            rows = conn.execute(
                "SELECT * FROM comprobantes ORDER BY created_at DESC LIMIT 100"
            ).fetchall()
            total_mes = conn.execute("""
                SELECT COALESCE(SUM(monto),0) FROM comprobantes
                WHERE estado='APROBADO'
                AND strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now')
            """).fetchone()[0]
            cant_mes = conn.execute("""
                SELECT COUNT(*) FROM comprobantes
                WHERE estado='APROBADO'
                AND strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now')
            """).fetchone()[0]
        else:
            sucursal = session["sucursal"]
            rows = conn.execute(
                "SELECT * FROM comprobantes WHERE sucursal=? ORDER BY created_at DESC LIMIT 100",
                (sucursal,)
            ).fetchall()
            total_mes = conn.execute("""
                SELECT COALESCE(SUM(monto),0) FROM comprobantes
                WHERE sucursal=? AND estado='APROBADO'
                AND strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now')
            """, (sucursal,)).fetchone()[0]
            cant_mes = conn.execute("""
                SELECT COUNT(*) FROM comprobantes
                WHERE sucursal=? AND estado='APROBADO'
                AND strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now')
            """, (sucursal,)).fetchone()[0]

    return render_template("dashboard.html",
        comprobantes=rows,
        total_mes=total_mes,
        cant_mes=cant_mes,
        usuario=session["usuario"],
        sucursal=session["sucursal"],
    )

@app.route("/emitir", methods=["GET", "POST"])
@login_required
def emitir():
    if request.method == "POST":
        tipo = request.form.get("tipo", "B")
        monto_total = float(request.form.get("monto", 0))

        if tipo == "A":
            neto = round(monto_total / 1.21, 2)
            iva  = round(monto_total - neto, 2)
        else:
            neto = monto_total
            iva  = 0.0

        data = {
            "tipo":     tipo,
            "nombre":   request.form.get("nombre"),
            "doc_tipo": int(request.form.get("doc_tipo", 99)),
            "doc_nro":  request.form.get("doc_nro", ""),
            "neto":     neto,
            "iva":      iva,
            "monto":    monto_total,
        }
        resultado = emitir_factura_arca(data)
        guardar_comprobante(resultado)

        if resultado["estado"] == "APROBADO":
            flash("Factura {} #{} emitida. CAE: {}".format(
                resultado["tipo"], resultado["numero"], resultado["cae"]), "success")
        else:
            flash("Error: {}".format(resultado["error"]), "error")

        return redirect(url_for("dashboard"))

    return render_template("emitir.html",
        sucursal=session["sucursal"],
        usuario=session["usuario"],
    )

@app.route("/masivo", methods=["GET", "POST"])
@login_required
def masivo():
    resultados = []
    if request.method == "POST":
        archivo = request.files.get("archivo")
        if not archivo:
            flash("Selecciona un archivo Excel o CSV", "error")
            return redirect(url_for("masivo"))

        nombre = archivo.filename.lower()
        filas = []

        if nombre.endswith(".xlsx") or nombre.endswith(".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    fila = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
                    filas.append(fila)
            except Exception as e:
                flash("Error al leer el Excel: {}".format(str(e)), "error")
                return redirect(url_for("masivo"))
        else:
            content = archivo.read().decode("utf-8")
            reader = csv.DictReader(io.StringIO(content))
            filas = [{k.lower().strip(): v for k, v in row.items()} for row in reader]

     for fila in filas:
            # Soporta tanto nombres nuevos (Excel Mundosim) como nombres viejos (CSV)
            tipo_raw = (fila.get("tipocomprobante") or fila.get("tipo") or "FB").strip().upper()
            # FA -> A, FB -> B
            tipo = tipo_raw.replace("F", "") if tipo_raw in ("FA", "FB") else tipo_raw

            nombre = (fila.get("razonsocial") or fila.get("nombre") or "").strip()
            monto_total = float(fila.get("monto") or 0)

            # Tipo de documento
            tipo_doc_raw = (fila.get("tipodocumento") or fila.get("doc_tipo") or "CF").strip().upper()
            if tipo_doc_raw in ("CUIT", "80"):
                doc_tipo = 80
            elif tipo_doc_raw in ("DNI", "96"):
                doc_tipo = 96
            else:
                doc_tipo = 99  # Consumidor Final

            doc_nro = (fila.get("nrodocumento") or fila.get("doc_nro") or "").strip()
            doc_nro = doc_nro.replace("-", "").replace(".", "")

            if tipo == "A":
                neto = round(monto_total / 1.21, 2)
                iva  = round(monto_total - neto, 2)
            else:
                neto = monto_total
                iva  = 0.0

            data = {
                "tipo":     tipo,
                "nombre":   nombre,
                "doc_tipo": doc_tipo,
                "doc_nro":  doc_nro,
                "neto":     neto,
                "iva":      iva,
                "monto":    monto_total,
            }
          
            resultado = emitir_factura_arca(data)
            guardar_comprobante(resultado)
            resultados.append(resultado)

    return render_template("masivo.html",
        sucursal=session["sucursal"],
        usuario=session["usuario"],
        resultados=resultados,
    )

def emitir_factura_arca(data):
    import random, string

    MODO_DEMO = not os.environ.get("CUIT")
    hoy = datetime.now().strftime("%Y-%m-%d")
    punto_venta = session.get("punto_venta") or 1
    tipo = data.get("tipo", "B")

    if MODO_DEMO:
        numero = random.randint(1000, 9999)
        cae = "".join(random.choices(string.digits, k=14))
        return {
            "numero":        numero,
            "tipo":          tipo,
            "punto_venta":   punto_venta,
            "sucursal":      session["sucursal"],
            "fecha":         hoy,
            "destinatario":  data["nombre"],
            "doc_tipo":      data["doc_tipo"],
            "doc_nro":       data["doc_nro"],
            "neto":          data["neto"],
            "iva":           data["iva"],
            "monto":         data["monto"],
            "cae":           cae,
            "cae_vto":       "20241231",
            "estado":        "APROBADO",
            "error":         None,
            "usuario":       session["usuario"],
        }

    try:
        from zeep import Client
        import subprocess, tempfile
        import xml.etree.ElementTree as ET
        from datetime import timedelta

        CUIT       = os.environ["CUIT"]
        CERT_PATH  = os.environ["CERT_PATH"]
        KEY_PATH   = os.environ["KEY_PATH"]
        PRODUCCION = os.environ.get("PRODUCCION", "false").lower() == "true"
        CONCEPTO   = int(os.environ.get("CONCEPTO", "2"))

        WSAA_URL = (
            "https://wsaa.afip.gov.ar/ws/services/LoginCms?wsdl"
            if PRODUCCION else
            "https://wsaahomo.afip.gov.ar/ws/services/LoginCms?wsdl"
        )
        WSFE_URL = (
            "https://servicios1.afip.gov.ar/wsfev1/service.asmx?WSDL"
            if PRODUCCION else
            "https://wswhomo.afip.gov.ar/wsfev1/service.asmx?WSDL"
        )

        cache_file = "/tmp/ticket_{}.json".format(CUIT)
        ticket = None
        if os.path.exists(cache_file):
            cached = json.loads(open(cache_file).read())
            if datetime.now() < datetime.fromisoformat(cached["expiry"]) - timedelta(minutes=10):
                ticket = cached

        if not ticket:
            now = datetime.utcnow()
            exp = now + timedelta(hours=12)
            xml_tra = """<?xml version="1.0" encoding="UTF-8"?>
<loginTicketRequest version="1.0">
  <header>
    <uniqueId>{}</uniqueId>
    <generationTime>{}</generationTime>
    <expirationTime>{}</expirationTime>
  </header>
  <service>wsfe</service>
</loginTicketRequest>""".format(
                int(now.timestamp()),
                now.strftime('%Y-%m-%dT%H:%M:%S'),
                exp.strftime('%Y-%m-%dT%H:%M:%S')
            ).encode()

            with tempfile.NamedTemporaryFile(suffix=".xml", delete=False, mode="wb") as f:
                f.write(xml_tra)
                tra_path = f.name
            with tempfile.NamedTemporaryFile(suffix=".cms", delete=False) as f:
                cms_path = f.name

            subprocess.run([
                "openssl", "smime", "-sign",
                "-in", tra_path, "-out", cms_path,
                "-signer", CERT_PATH, "-inkey", KEY_PATH,
                "-outform", "PEM", "-nodetach"
            ], check=True, capture_output=True)

            cms_b64 = "".join(
                l for l in open(cms_path).read().splitlines()
                if not l.startswith("-----")
            )
            os.unlink(tra_path)
            os.unlink(cms_path)

            wsaa = Client(WSAA_URL)
            resp_xml = wsaa.service.loginCms(in0=cms_b64)
            root = ET.fromstring(resp_xml)

            ticket = {
                "token":  root.find(".//token").text,
                "sign":   root.find(".//sign").text,
                "expiry": exp.isoformat(),
            }
            open(cache_file, "w").write(json.dumps(ticket))

        auth = {"Token": ticket["token"], "Sign": ticket["sign"], "Cuit": int(CUIT)}
        tipo_codigo = TIPOS_COMPROBANTE[tipo]["codigo"]

        wsfe = Client(WSFE_URL)
        ultimo = wsfe.service.FECompUltimoAutorizado(
            Auth=auth, PtoVta=punto_venta, CbteTipo=tipo_codigo
        ).CbteNro
        numero = ultimo + 1

        hoy_arca = datetime.now().strftime("%Y%m%d")
        neto  = round(data["neto"], 2)
        iva   = round(data["iva"], 2)
        total = round(data["monto"], 2)

        alicuotas = []
        if tipo == "A" and iva > 0:
            alicuotas = [{"Id": 5, "BaseImp": neto, "Importe": iva}]

        detalle = {
            "Concepto":     CONCEPTO,
            "DocTipo":      data["doc_tipo"],
            "DocNro":       int(data["doc_nro"]) if data["doc_nro"] else 0,
            "CbteDesde":    numero,
            "CbteHasta":    numero,
            "CbteFch":      hoy_arca,
            "ImpTotal":     total,
            "ImpTotConc":   0,
            "ImpNeto":      neto,
            "ImpOpEx":      0,
            "ImpIVA":       iva,
            "ImpTrib":      0,
            "FchServDesde": hoy_arca if CONCEPTO in (2, 3) else None,
            "FchServHasta": hoy_arca if CONCEPTO in (2, 3) else None,
            "FchVtoPago":   hoy_arca if CONCEPTO in (2, 3) else None,
            "MonId":        "PES",
            "MonCotiz":     1,
        }

        if alicuotas:
            detalle["Iva"] = {"AlicIva": alicuotas}

        resp = wsfe.service.FECAESolicitar(Auth=auth, FeCAEReq={
            "FeCabReq": {"CantReg": 1, "PtoVta": punto_venta, "CbteTipo": tipo_codigo},
            "FeDetReq": {"FECAEDetRequest": [detalle]},
        })

        det = resp.FeDetResp.FECAEDetResponse[0]
        if det.Resultado == "A":
            return {
                "numero": numero, "tipo": tipo,
                "punto_venta": punto_venta,
                "sucursal": session["sucursal"], "fecha": hoy,
                "destinatario": data["nombre"], "doc_tipo": data["doc_tipo"],
                "doc_nro": data["doc_nro"], "neto": neto, "iva": iva, "monto": total,
                "cae": det.CAE, "cae_vto": det.CAEFchVto,
                "estado": "APROBADO", "error": None,
                "usuario": session["usuario"],
            }
        else:
            errores = []
            if det.Observaciones:
                for o in det.Observaciones.Obs:
                    errores.append("[{}] {}".format(o.Code, o.Msg))
            raise Exception("; ".join(errores))

    except Exception as e:
        return {
            "numero": None, "tipo": tipo,
            "punto_venta": punto_venta,
            "sucursal": session["sucursal"], "fecha": hoy,
            "destinatario": data["nombre"], "doc_tipo": data.get("doc_tipo"),
            "doc_nro": data.get("doc_nro"), "neto": data.get("neto"),
            "iva": data.get("iva"), "monto": data["monto"],
            "cae": None, "cae_vto": None,
            "estado": "RECHAZADO", "error": str(e),
            "usuario": session["usuario"],
        }

def guardar_comprobante(r):
    with get_db() as conn:
        conn.execute("""
            INSERT INTO comprobantes
            (numero, tipo, punto_venta, sucursal, fecha, destinatario, doc_tipo, doc_nro,
             neto, iva, monto, cae, cae_vto, estado, error, usuario)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            r.get("numero"), r.get("tipo"), r.get("punto_venta"), r.get("sucursal"),
            r.get("fecha"), r.get("destinatario"), r.get("doc_tipo"), r.get("doc_nro"),
            r.get("neto"), r.get("iva"), r.get("monto"), r.get("cae"),
            r.get("cae_vto"), r.get("estado"), r.get("error"), r.get("usuario"),
        ))
        conn.commit()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
